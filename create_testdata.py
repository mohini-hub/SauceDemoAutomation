"""
Run this script to create LoginData.xlsx test data file.

Steps:
1. pip install openpyxl
2. python3 create_testdata.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_login_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LoginData"

    # Header row styling
    headers = ["Username", "Password", "ExpectedResult"]
    header_fill = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # SauceDemo test users data
    test_data = [
        ["standard_user",   "secret_sauce", "pass"],  # Valid user - should pass
        ["locked_out_user", "secret_sauce", "fail"],  # Locked user - should fail
        ["problem_user",    "secret_sauce", "pass"],  # Problem user - login works
        ["standard_user",   "wrong_pass",   "fail"],  # Wrong password - should fail
        ["",                "secret_sauce", "fail"],  # Empty username - should fail
    ]

    # Alternating row colors
    row_colors = ["EBF3FF", "FFFFFF"]
    for row_idx, row_data in enumerate(test_data, 2):
        fill_color = row_colors[row_idx % 2]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18

    # Save file
    output_path = "src/test/resources/testdata/LoginData.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"LoginData.xlsx created at: {output_path}")

if __name__ == "__main__":
    create_login_data()
